#!Python3
# Created by Thomas 'Lily' Lilienthal OCT.2024 - Deschutes County, Oregon - v1
'''
This will take a passed cybertip .json file and parse out two excel files:
    -_reportID_Parsed_IPs.xlsx
    -_reportID_Parsed_Media.xlsx

requirements:
requests, openpyxl
'''
import sys, os, openpyxl, json
from openpyxl.utils import get_column_letter
from ARIN_Query import queryIP
from datetime import datetime

def formatWorksheet(worksheet, headerRow):
    # Make row headerRow the header and apply filters
    worksheet.auto_filter.ref = f"A{headerRow}:{get_column_letter(worksheet.max_column)}{headerRow}"

    # Freeze the top headerRow rows
    worksheet.freeze_panes = f"A{headerRow+1}"

    # Adjust column widths to fit content
    for col in worksheet.columns:
        maxLength = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if cell.value:
                    maxLength = max(maxLength, len(str(cell.value)))
            except:
                pass
        adjustedWidth = (maxLength + 2)
        worksheet.column_dimensions[column].width = adjustedWidth

def ipFile(inFile, outputPath, key):
    # Arin.net timesout after 100 queries, so to limit the possibility of that, we cache IPs
    # Also, create a counter for new IPs with a timer in case we hit 90 in a minute
    cachedIPs = {}

    # convert the input to a Python dictionary using the json.loads() function
    with open(inFile, 'r', encoding='utf-8-sig') as file: data = json.load(file)

    # Most of the good stuff is in the reportedInfo object
    reportedInfo = data["reportedInformation"] 

    # Nested list parsedIPs. Create variable and generate the headers
    parsedIPs = [['IP Address', 'Datetime', 'ISP', 'Source / TIP Filename', 'Type']]

    # Get the reported person IP addresses
    reportedPeople = reportedInfo['reportedPeople']

    for reportedPerson in reportedPeople['reportedPersons']:
        # Get email if available.
        email = 'Suspect Account'
        # Parse through the sourceInformation section where the Login and other IPs are located
        if reportedPerson['sourceInformation']:
            sourceInformation = reportedPerson['sourceInformation']['sourceCaptures']
            for a in sourceInformation:
                if 'IP' in a['captureType']:
                    ip = a['value']
                    dateTime = a['dateTime']
                    if dateTime: formattedDateTime = f'{datetime.strptime(dateTime, '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d %H:%M:%S')} UTC'
                    else: formattedDateTime = dateTime
                    eventName = a['eventName']
                    
                    # Add to list
                    parsedIPs.append([ip, formattedDateTime, '', email, eventName])
                    
    # Get the uploaded file IP addresses
    uploadedFiles = reportedInfo['uploadedFiles']
    for uploadedFile in uploadedFiles['uploadedFiles']:
        # Grab filenames first. 
        filename = uploadedFile['filename']

        # Grab IP info and eventName
        if uploadedFile['sourceInformation']: 
            for a in uploadedFile['sourceInformation']['sourceCaptures']:
                if 'IP' in a['captureType']:
                    ip = a['value']
                    dateTime = a['dateTime']
                    if dateTime: formattedDateTime = f'{datetime.strptime(dateTime, '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d %H:%M:%S')} UTC'
                    else: formattedDateTime = dateTime
                    eventName = a['eventName']
                    
                    # Add to list
                    parsedIPs.append([ip, formattedDateTime, '', filename, eventName])
                    
    # Now get all the IP addresses and pass them for queryIPs. Probably a better way to do this but /shurg
    # Make list of only IP addresses for queryIP. It will do its own dedupe
    ipList = []
    for a in parsedIPs: 
        if a[0] != 'IP Address': ipList.append(a[0])
    
    # New dictionary with return
    ispDict = queryIP(ipList, key)

    # int iterate is created to move throught the nested lists for navigation of updating ISP info
    iterate = 0

    # Loop through, when an IP address matches in the dictionary, add the ISP info
    for a in parsedIPs:
        if a[0] in ispDict: parsedIPs[iterate][2] = ispDict[a[0]]
        iterate+=1

    # Make new excel files from parsed CT IP addresses.
    newFile = f'''_CT{data['reportId']}_Parsed_IPs.xlsx'''
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for row in parsedIPs: worksheet.append(row)
    formatWorksheet(worksheet, 1)
    workbook.save(f'''{outputPath}/{newFile}''')
    workbook.close()

    print(f'''A total of {len(ipList)} have been parsed from the cybertip and queried. {newFile} has been created!''')

def suspectInfo(data):
    # CTs appear to have no consistency in how they store key:value pairs. Sometimes, its NULL, or doesnt exist at all
    # This just handles potential NOT NULL errors later
    name = ''
    screenName = 'None'
    id = 'None'
    email = 'None'
    dob = 'None'
    number = 'None'

    if data['firstName']: name += data['firstName'] 
    if data['middleName']: name += f' {data['middleName']}' 
    if data['lastName']: name += f' {data['lastName']}'
    if name == '': name = "None"
    if data['screenName']: screenName = data['screenName']['value']
    if data['espUserId']: id = data['espUserId']
    if data['dateOfBirth']: dob = data['dateOfBirth']
    if data['emails']:
        for a in data['emails']['emails']: 
            if email == 'None': email = ''
            else: email += ', '
            email += f'{a['value']} ({a['verified']})'
    if data['phones']:
        for a in data['phones']['phones']:
            if number == 'None': number =''
            else: number += ', '
            number += f'{a['value']} ({a['verified']})'
    
    return([name.strip(), screenName, id, dob, email, number])

def mediaFile(inFile, outputPath):
    # Parses the json inFile cybertip, and creates the excel file the the media parsed at the outputPath

    # convert the input to a Python dictionary using the json.loads() function
    with open(inFile, 'r', encoding='utf-8-sig') as file: data = json.load(file)

    # Most of the good stuff is in the reportedInfo object
    reportedInfo = data["reportedInformation"] 

    # Get the "reportingEsp" object and extract the espName
    reportingEsp = reportedInfo["reportingEsp"]
    espName = reportingEsp["espName"]

    incidentSummary = reportedInfo['incidentSummary']
    incidentType = incidentSummary['incidentType']
    incidentDateTime = incidentSummary['incidentDateTime']
    incidentDateTime = f'{datetime.strptime(incidentDateTime, '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d %H:%M:%S')} UTC'

    # Get suspect information into a list
    suspectAccount = suspectInfo(reportedInfo['reportedPeople']['reportedPersons'][0])

    # Nested list parsedMedia. Create variable and generate the headers
    parsedMedia = []

    # Get the uploaded file IP addresses
    uploadedFiles = reportedInfo['uploadedFiles']
    totalUploads = 0
    parsedMedia.append([f'Cybertip={data["reportId"]}', f'Reporting ESP={espName}', f'Incident Type={incidentType}', f'Incident Datetime={incidentDateTime}'])
    parsedMedia.append([f'Suspect Name={suspectAccount[0]}', f'Suspect User Name={suspectAccount[1]}', f'Suspect ID={suspectAccount[2]}', f'Suspect DOB={suspectAccount[3]}', f'Suspect Emails={suspectAccount[4]}', f'Suspect Phone Numbers={suspectAccount[5]}'])
    parsedMedia.append(['Tip Filename', 'Original Filename', 'Event', 'IP Address', 'Datetime', 'Classification', 'Viewed by ESP?', 'NCMEC Tags', 'Hash', 'Additional Information'])
    

    for uploadedFile in uploadedFiles['uploadedFiles']:
        totalUploads+=1

        # Grab filenames first
        filename = uploadedFile['filename']
        originalFilename = uploadedFile['originalFilename']
        industryClassification = uploadedFile['industryClassification']
        viewedByEsp = str(uploadedFile['viewedByEsp'])
        if viewedByEsp == 'None': viewedByEsp = 'No Information Provided'
        verificationHash = uploadedFile['verificationHash']
        ncmecTags = 'None'
        formattedDateTime = 'Unknown'
        additionalInformation = 'None'
        event = 'Unknown'
        ipAddress = 'Unknown'
        
        if uploadedFile['sourceInformation']: 
            for a in uploadedFile['sourceInformation']['sourceCaptures']:
                if 'IP' in a['captureType']:
                    dateTime = a['dateTime']
                    ipAddress = a['value']
                    event = a['eventName']
                    formattedDateTime = f'{datetime.strptime(dateTime, '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d %H:%M:%S')} UTC'
        if uploadedFile['ncmecTags']: 
            ncmecTags = ''
            for a in uploadedFile['ncmecTags']['groups']:
                if ncmecTags != '': ncmecTags += ', '
                ncmecTags += f'''{a['tags'][0]['value']}'''
        if uploadedFile['additionalInformations']: additionalInformation = uploadedFile['additionalInformations'][0]['value']

                    
        # Add to list
        parsedMedia.append([filename, originalFilename, event, ipAddress, formattedDateTime, industryClassification, viewedByEsp, ncmecTags, verificationHash, additionalInformation])
    
    parsedMedia[0].append(f'Total Uploads={totalUploads}')
    
    # Make new file named the tip.
    newFile = f'''_CT{data['reportId']}_Parsed_Multimedia.xlsx'''
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    for row in parsedMedia: worksheet.append(row)
    formatWorksheet(worksheet, 3)
    workbook.save(f'''{outputPath}/{newFile}''')
    workbook.close()

    print(f'''A total of {totalUploads} media files have been parse from the cybertip. {newFile} has been created!''')
    
def main():
    # Check if file path was passed (as it would be when dragging and dropping or via command-line)
    if len(sys.argv) < 2:
        print("Error: No file provided.")
        return

    # Get the json file path from command-line arguments
    jsonFile = os.path.normpath(sys.argv[1])

    # Ensure the file has an .json extension
    if not jsonFile.endswith('.json'):
        print("Error: Provided file is not a json document.")
        return

    # Set the output path in the same directory as the input json file
    outputPath = os.path.dirname(jsonFile)
    
    # Get the directory of the current script
    scriptDirectory = os.path.dirname(os.path.abspath(__file__))
    
    # Path to 'apiKey.txt' in the same directory as the script
    apiFilePath = os.path.join(scriptDirectory, 'apiKey.txt')
    
    # Check if 'apiKey.txt' exists
    if os.path.exists(apiFilePath):
        with open(apiFilePath, 'r') as apiFile: apiKey = apiFile.readline().strip()
        print('API Key located. We will use this for arin.net queries.')
    else:
        apiKey = input('Please enter an API key if you have one. This script will save that for all future queries. If you don\'t have one, just click enter: ')
        if len(apiKey) == 43:
            with open(apiFilePath, 'w', newline='') as apiFile: apiFile.write(apiKey)
            print('API Key accepted and saved. We will use this for arin.net queries.')
        elif apiKey == '': print('You did not enter an API Key. This will still work but will be significantly slower if you have a lot of IP addresses.')
        else: 
            apiKey = ''
            print('Your API Key appears invalid. If you think this is a mistake, please close and rerun the script.')

    # Parse and query IP addresses to excel file.
    ipFile(jsonFile, outputPath, apiKey)

    # Parse media files to excel file.
    mediaFile(jsonFile, outputPath)

    done = input('Processing complete! Press enter to exit.')

if __name__ == "__main__":
    main()